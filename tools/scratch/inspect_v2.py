import openpyxl, warnings
from collections import Counter
warnings.filterwarnings("ignore")

wb = openpyxl.load_workbook("docs/org-notes-compilation/Team List Billed v2.xlsx", data_only=True, read_only=True)
s = wb["Trupti -Updated Employee Master"]
rows = list(s.iter_rows(values_only=True))
print("v2 header row:")
for i, h in enumerate(rows[0]):
    print(f"  [{i}] {h!r}")
print()
print("first 6 data rows, cols 8..12:")
for r in rows[1:7]:
    print("  ", r[8:13])

# Headers v2 = 13 cols. Unnamed are at idx 9 and 10.
print()
print("col idx 9 distinct (top 20):")
def safe(r, i): return r[i] if i < len(r) else None
c9 = Counter([safe(r,9) for r in rows[1:] if safe(r,9) is not None])
for k, v in c9.most_common(20):
    print(f"  {v:4d}  {k!r}")
print(f"  [{sum(1 for r in rows[1:] if safe(r,9) is None)} blanks]")
print()
print("col idx 10 distinct (top 20):")
c10 = Counter([safe(r,10) for r in rows[1:] if safe(r,10) is not None])
for k, v in c10.most_common(20):
    print(f"  {v:4d}  {k!r}")
print(f"  [{sum(1 for r in rows[1:] if safe(r,10) is None)} blanks]")

print()
print("=== v1 Reliance Counterpart distribution (top 20) ===")
wb1 = openpyxl.load_workbook("docs/org-notes-compilation/Team List Billed.xlsx", data_only=True, read_only=True)
s1 = wb1["Trupti -Updated Employee Master"]
r1 = list(s1.iter_rows(values_only=True))
print("v1 header:")
for i, h in enumerate(r1[0]):
    print(f"  [{i}] {h!r}")
c1 = Counter([r[9] for r in r1[1:] if r[9] is not None])
for k, v in c1.most_common(20):
    print(f"  {v:4d}  {k!r}")
print(f"  [{sum(1 for r in r1[1:] if r[9] is None)} blanks]")
