#!/usr/bin/env python3
"""Build organisation/data.json — the single source of truth for every
org-wide number on the Fynd × Reliance Retail register.

Run from repo root:
    python3 tools/build_org_data.py

Reads three sources, all local-only and gitignored:
  1. docs/org-notes-compilation/Team List Billed v2.xlsx
       sheet: 'Trupti -Updated Employee Master'
       1,056 non-blank rows · billed roster · authoritative for project
       assignment, billing entity, tier rollups, headcount totals.
  2. ~/Documents/work/engineering-os/docs/onboarding/keka_roster_full.csv
       1,095 rows · keka HRIS export · authoritative for the reporting
       tree (Reporting To chain) and per-person metadata (CXO, Date
       Joined, Location). Round-3 addition.
  3. ~/Documents/work/engineering-os/backend/data/keka/employees.csv
       4-column employee_id → photo_path map · authoritative for which
       employee has a CDN photo URL.

Reads two committed config files:
  · organisation/mappings.json   — xlsx → display name + route + tier;
                                   tracks; placeholder names; tier model.
  · organisation/leadership.json — founders + C-suite + project leads
                                   (hand-edited).

Writes one committed artefact:
  · organisation/data.json — consumed by all 4 /organisation pages on load.
                             Top-level: meta + facets + tiers + leaders +
                             tree + pivot + rows.

Prints a build summary at end including an UNMAPPED REPORT — any xlsx
Project / Track value not found in mappings.json. Resolve by adding the
entry to mappings.json before committing.

For the full lifecycle (sources, schema details, page-by-page render
contract, verification protocol, common-edit recipes), read:
    docs/organisation-workflow.md

Build-round history (per docs/organisation-spec.md §9):
  · 2026-05-03 round 1 — pinned to v2.xlsx; dropped Reliance Counterpart
    field (v2 source removed the column).
  · 2026-05-03 round 2 — emits 'tier' per row + per pivot, 'topTracks'
    (top-5 by count + moreTracks rest count), 'tiers' rollup, 'leaders'
    block from leadership.json, end-of-build unmapped report.
  · 2026-05-03 round 3 — ingests the richer keka_roster_full.csv; emits
    'tree' block (nodes + rootNames + children); enriches data.rows with
    reportingTo / cxo / dateJoined / location where Employee Number joins.
  · 2026-05-03 round 3 follow-up — placeholderNames filter (drops org-chart
    placeholder seats from the tree, reparents reports up the chain);
    billed-only tree filter (only people whose Display Name appears in
    data.rows are rendered in the organogram; same chain-reparent logic).
"""
from __future__ import annotations
import csv, json, os, sys, warnings
from collections import Counter, defaultdict
from pathlib import Path

warnings.filterwarnings("ignore")
import openpyxl  # type: ignore

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "docs/org-notes-compilation/Team List Billed v2.xlsx"
SHEET = "Trupti -Updated Employee Master"
KEKA_CSV = Path("/Users/kushanshah/Documents/work/engineering-os/backend/data/keka/employees.csv")
KEKA_ROSTER_FULL = Path("/Users/kushanshah/Documents/work/engineering-os/docs/onboarding/keka_roster_full.csv")
OUT = ROOT / "organisation/data.json"
MAPPINGS = ROOT / "organisation/mappings.json"
LEADERSHIP = ROOT / "organisation/leadership.json"

CDN_BASE = "https://socialassets.impetusz0.de/rrl-portfolio/assets/keka-photos"

# Columns to drop entirely before writing data.json (spec § 5.3).
SENSITIVE_COLS = {"Average Rating 25-26", "Average Rating 24-25", "Current TCTC"}


def load_mappings() -> tuple[dict, dict, dict, dict, set[str], list[dict]]:
    """Read organisation/mappings.json into (projects, tracks, kinds, tiers, placeholders, splits).

    Track keys are lowercased + trimmed for case-insensitive matching against
    xlsx values that drift in capitalisation. Placeholders is a set of names
    in the keka roster that are org-chart placeholder seats (open hires)
    rather than real people — these are excluded from the organogram tree.
    Splits is a list of {fromProject, track, newProject, route, tier} entries
    for tracks that should be promoted to their own virtual project on the
    /organisation page (e.g., Retail Vista is filed under UCP & Marketing OS
    in the xlsx but presented as a standalone card).
    """
    with MAPPINGS.open() as f:
        m = json.load(f)
    projects = m.get("projects", {})
    tracks = {k.strip().lower(): v for k, v in m.get("tracks", {}).items()}
    kinds = m.get("kinds", {})
    tiers = m.get("tiers", {})
    placeholders = set(m.get("placeholderNames", []))
    splits = m.get("splitTracksAsProjects", [])
    return projects, tracks, kinds, tiers, placeholders, splits


def load_leadership() -> dict:
    """Read organisation/leadership.json. Returns empty skeleton if missing."""
    if not LEADERSHIP.exists():
        return {"founders": [], "cSuite": [], "projectLeads": {}}
    with LEADERSHIP.open() as f:
        return json.load(f)


def load_full_roster() -> tuple[list[dict], dict[str, dict]]:
    """Read the richer keka roster (manager + cxo + location + dateJoined).

    Returns (rows, by_emp_id) where by_emp_id maps Employee Number → row dict.
    Used for two things:
      1. Enrich data.rows with reportingTo / cxo / dateJoined / location.
      2. Build the organogram tree (uses ALL rows here, including founders +
         non-billed staff that are not in the billed xlsx).
    """
    if not KEKA_ROSTER_FULL.exists():
        return [], {}
    rows = []
    with KEKA_ROSTER_FULL.open() as f:
        for r in csv.DictReader(f):
            rows.append(r)
    by_emp = {(r.get("Employee Number") or "").strip(): r for r in rows if r.get("Employee Number")}
    return rows, by_emp


def build_tree(full_rows: list[dict], photo_map: dict[str, str],
               placeholders: set[str], visible_names: set[str]) -> dict:
    """Build the organogram tree from the full keka roster, filtered to billed.

    Per round-3 follow-up direction: the organogram only displays members of
    the billed sheet (via Display Name match). The keka roster is consulted
    for reporting lines (Reporting To) and per-node metadata (CXO, Date
    Joined, Location), but a row is rendered only if its Display Name is in
    `visible_names`.

    Each rendered node carries: id, name, title, dept, location, photo,
    reportingTo, cxo, project, dateJoined. Adjacency = parent name → [child
    names]. Roots = self-reporting nodes (founders) that are also visible.

    Two filter passes:
      1. Placeholders (mappings.json > placeholderNames) — open hires + dummy
         seats. Resolved to their first non-placeholder ancestor.
      2. Non-billed (any keka row whose Display Name isn't in visible_names) —
         e.g., Sreeraman / SMG who is in keka but not billed. Resolved up the
         keka reporting chain to the first visible ancestor.

    Both filters use the same chain-resolution mechanism. When a row is
    dropped, its reports are reparented up the chain so nobody is orphaned.
    If the chain dies inside the dropped set (no visible ancestor reachable),
    the descendant becomes a root.
    """
    nodes: list[dict] = []
    by_name: dict[str, dict] = {}
    children: dict[str, list[str]] = defaultdict(list)
    root_names: list[str] = []
    dropped_placeholder: list[str] = []
    dropped_nonbilled: list[str] = []

    # Pre-pass: build name → reportingTo map from the full keka roster.
    rt_by_name: dict[str, str] = {}
    for r in full_rows:
        n = (r.get("Display Name") or "").strip()
        if n:
            rt_by_name[n] = (r.get("Reporting To") or "").strip()

    # The drop set = placeholders ∪ (keka names not in billed). Anyone in
    # this set is excluded from the tree; their reports reparent up the chain.
    drop_set: set[str] = set(placeholders) | (set(rt_by_name) - visible_names)

    def resolve_visible(start: str) -> str:
        """Walk Reporting To chain from `start` until we hit a visible
        non-dropped ancestor or terminate (cycle/missing). Returns the
        resolved name or '' if no visible ancestor exists.
        """
        seen = {start}
        cur = rt_by_name.get(start, "")
        while cur in drop_set and cur not in seen:
            seen.add(cur)
            cur = rt_by_name.get(cur, "")
        return cur if cur in visible_names else ""

    # Cache: each dropped name resolves to the same visible ancestor.
    drop_resolves: dict[str, str] = {d: resolve_visible(d) for d in drop_set if d in rt_by_name}

    for r in full_rows:
        name = (r.get("Display Name") or "").strip()
        if not name:
            continue
        if name in placeholders:
            dropped_placeholder.append(name)
            continue
        if name not in visible_names:
            dropped_nonbilled.append(name)
            continue
        emp_id = (r.get("Employee Number") or "").strip()
        rt = (r.get("Reporting To") or "").strip()
        # Reparent through dropped (placeholder OR non-billed) chain.
        if rt in drop_set:
            rt = drop_resolves.get(rt, "")
        node = {
            "id": emp_id,
            "name": name,
            "title": (r.get("Job Title") or "").strip() or None,
            "dept": (r.get("Department") or "").strip() or None,
            "location": (r.get("Location") or "").strip() or None,
            "photo": photo_map.get(emp_id),
            "reportingTo": rt or None,
            "cxo": (r.get("CXO's") or "").strip() or None,
            "project": (r.get("PROJECT") or "").strip() or None,
            "dateJoined": (r.get("Date Joined") or "").strip() or None,
        }
        nodes.append(node)
        by_name.setdefault(name, node)
        if rt and rt != name:
            children[rt].append(name)
        else:
            # Self-reporting or blank Reporting To = root candidate.
            root_names.append(name)

    # Children sort: by direct-report subtree size (biggest first), then name.
    def subtree_size(n: str, seen: set[str] | None = None) -> int:
        if seen is None:
            seen = set()
        if n in seen:
            return 0
        seen.add(n)
        return 1 + sum(subtree_size(c, seen) for c in children.get(n, []))
    for parent, kids in children.items():
        kids.sort(key=lambda n: (-subtree_size(n), n))

    # Direct reports count per node.
    for node in nodes:
        node["directReports"] = len(children.get(node["name"], []))

    return {
        "nodes": nodes,
        "rootNames": sorted(set(root_names)),
        "children": dict(children),
        "droppedPlaceholder": sorted(set(dropped_placeholder)),
        "droppedNonBilled": sorted(set(dropped_nonbilled)),
    }


def load_keka_photo_map() -> dict[str, str]:
    """Map xlsx Employee Number (string of int) -> CDN photo URL.

    Keka photos live at gs://impetus-socialpilot/rrl-portfolio/assets/keka-photos/<id>.<ext>.
    Filename derived from the photo_path column ('photos/C483.jpeg' -> 'C483.jpeg').
    """
    m: dict[str, str] = {}
    with KEKA_CSV.open() as f:
        for row in csv.DictReader(f):
            pp = (row.get("photo_path") or "").strip()
            if not pp:
                continue
            fname = os.path.basename(pp)
            m[row["employee_id"].strip()] = f"{CDN_BASE}/{fname}"
    return m


def normalize_emp_id(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def main() -> int:
    if not XLSX.exists():
        print(f"ERR: {XLSX} not found", file=sys.stderr)
        return 1
    if not KEKA_CSV.exists():
        print(f"ERR: {KEKA_CSV} not found", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    s = wb[SHEET]

    headers = [s.cell(1, c).value for c in range(1, s.max_column + 1)]
    keep_idx = [i for i, h in enumerate(headers) if h not in SENSITIVE_COLS]
    keep_headers = [headers[i] for i in keep_idx]

    photo_map = load_keka_photo_map()
    PROJECT_MAP, TRACK_MAP, KIND_MAP, TIER_MAP, PLACEHOLDERS, SPLITS = load_mappings()
    # Lookup: (fromProject displayName, track displayName) → split spec.
    # Used to re-assign rows + pivot entries after the main loop.
    SPLIT_MAP = {(s["fromProject"], s["track"]): s for s in SPLITS}
    LEADERSHIP_DATA = load_leadership()
    full_rows, full_by_emp = load_full_roster()

    # Track unmapped values for the end-of-build report.
    unmapped_projects: set[str] = set()
    unmapped_tracks: set[str] = set()
    untiered_projects: set[str] = set()

    rows: list[dict] = []
    photo_hits = 0
    for r in range(2, s.max_row + 1):
        raw = [s.cell(r, c).value for c in range(1, s.max_column + 1)]
        if raw[0] is None:
            continue
        rec = {keep_headers[j]: raw[i] for j, i in enumerate(keep_idx)}
        emp_id = normalize_emp_id(rec.get("Employee Number"))
        rec["Employee Number"] = emp_id

        project_raw = (rec.get("Project") or "").strip() or None
        track_raw = (rec.get("Track") or "").strip() or None

        proj_meta = PROJECT_MAP.get(project_raw, {}) if project_raw else {}
        track_meta = TRACK_MAP.get(track_raw.lower(), {}) if track_raw else {}

        # Track drift: a non-empty xlsx value with no mapping entry.
        if project_raw and project_raw not in PROJECT_MAP:
            unmapped_projects.add(project_raw)
        if track_raw and track_raw.lower() not in TRACK_MAP:
            unmapped_tracks.add(track_raw)
        # Tier coverage check (every mapped project should declare a tier).
        if project_raw and project_raw in PROJECT_MAP and not proj_meta.get("tier"):
            untiered_projects.add(project_raw)

        project_display = proj_meta.get("displayName") or project_raw
        track_display = track_meta.get("displayName") or track_raw
        tier = proj_meta.get("tier")

        # Normalise the Dedicated/Platform field (e.g., 'Platform' -> 'Shared').
        kind_raw = rec.get("Dedicated/ Platform")
        kind = KIND_MAP.get(kind_raw, kind_raw) if kind_raw else None

        photo_url = photo_map.get(emp_id)
        if photo_url:
            photo_hits += 1

        # Enrich with full-roster fields where available (CXO, location, etc).
        full = full_by_emp.get(emp_id, {})
        cxo = (full.get("CXO's") or "").strip() or None
        location = (full.get("Location") or "").strip() or None
        date_joined = (full.get("Date Joined") or "").strip() or None
        reporting_to = (full.get("Reporting To") or "").strip() or None

        rows.append({
            "id": emp_id,
            "name": rec.get("Display Name"),
            "department": rec.get("Department"),
            "title": rec.get("Job Title"),
            "manager": rec.get("Reporting Manager"),
            "billing": rec.get("Billing Entity"),
            "project": project_display,
            "projectRoute": proj_meta.get("route"),
            "tier": tier,
            "track": track_display,
            "trackRoute": track_meta.get("route"),
            "trackChips": track_meta.get("chips") or [],
            "kind": kind,
            "photo": photo_url,
            "cxo": cxo,
            "location": location,
            "dateJoined": date_joined,
            "reportingTo": reporting_to,
        })

    # Apply splitTracksAsProjects: re-assign matching rows from the parent
    # project to the new virtual project, so the directory drill-down works
    # and the pivot construction below produces a separate entry. Tracks set
    # to None in the rewritten rows so the new project's pivot has a single
    # implicit "self" track equal to the project (no sub-rows needed).
    for r in rows:
        spec = SPLIT_MAP.get((r.get("project"), r.get("track")))
        if spec is None:
            continue
        r["project"] = spec["newProject"]
        r["projectRoute"] = spec.get("route")
        r["tier"] = spec.get("tier") or r["tier"]
        # Promote the track to be the single track of the new project. We keep
        # the track row visible (one line) so the card still has content.

    # Summary stats for the hero block (live numbers, per spec § 5.4).
    total = len(rows)
    on_rrl = sum(1 for x in rows if (x["billing"] or "").upper() == "RRL")
    eng = sum(1 for x in rows if (x["department"] or "").lower() == "engineering")
    # "Product Builders" = the broad set Apex sees on the homepage §09 — not just
    # engineers, but every department that ships product. Definition kept in sync
    # with index.html so /organisation tile and homepage tile reconcile.
    PRODUCT_BUILDER_DEPTS = {
        "engineering", "product", "design", "data analysis",
        "data science", "data annotator", "quality assurance",
        "research", "design research", "fashion design",
    }
    builders = sum(1 for x in rows if (x["department"] or "").lower() in PRODUCT_BUILDER_DEPTS)
    # Pivot uses display names directly (rows already carry the normalised strings).
    # Track routes are keyed by (project, track) because two distinct xlsx Track
    # values can normalise to the same display name in different projects (e.g.,
    # "Plan: Cortex" in Impetus and standalone "Cortex" in Granary both display
    # as "Cortex" but have different routes).
    pivot: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    project_routes: dict[str, str | None] = {}
    project_tiers: dict[str, str | None] = {}
    track_routes: dict[tuple[str, str], str | None] = {}
    track_chips: dict[tuple[str, str], list] = {}
    for x in rows:
        p = x["project"] or "—"
        t = x["track"] or "—"
        pivot[p][t] += 1
        project_routes[p] = x["projectRoute"]
        project_tiers[p] = x["tier"]
        track_routes[(p, t)] = x["trackRoute"]
        track_chips[(p, t)] = x["trackChips"]

    pivot_out = []
    for proj, tracks in sorted(pivot.items(), key=lambda kv: -sum(kv[1].values())):
        track_rows = sorted(tracks.items(), key=lambda kv: -kv[1])
        full_tracks = [{
            "track": t,
            "trackRoute": track_routes.get((proj, t)),
            "trackChips": track_chips.get((proj, t), []),
            "count": c,
        } for t, c in track_rows]
        pivot_out.append({
            "project": proj,
            "projectRoute": project_routes.get(proj),
            "tier": project_tiers.get(proj),
            "total": sum(tracks.values()),
            "tracks": full_tracks,
            # topTracks = first 5 by count; rest live in tracks for drill-down.
            "topTracks": full_tracks[:5],
            "moreTracks": max(0, len(full_tracks) - 5),
        })

    # Filter facets (small, useful).
    def facets(field: str) -> list[str]:
        vals = sorted({(x[field] or "—") for x in rows})
        return vals

    # Tier rollup — sum people per tier so the Overview can render headcounts.
    tier_totals: dict[str, int] = defaultdict(int)
    for r in rows:
        if r.get("tier"):
            tier_totals[r["tier"]] += 1

    out = {
        "meta": {
            "source": "docs/org-notes-compilation/Team List Billed v2.xlsx · sheet: Trupti -Updated Employee Master",
            "generated": "build via tools/build_org_data.py",
            "totals": {
                "rows": total,
                "onRRL": on_rrl,
                "engineering": eng,
                "engineeringPct": round(eng / total * 100) if total else 0,
                "productBuilders": builders,
                "productBuildersPct": round(builders / total * 100) if total else 0,
                "photoHits": photo_hits,
                "photoMisses": total - photo_hits,
            },
        },
        "facets": {
            "project": facets("project"),
            "track": facets("track"),
            "department": facets("department"),
            "billing": facets("billing"),
            "kind": facets("kind"),
        },
        "tiers": [
            {"key": k, "label": v.get("label"), "order": v.get("order"), "total": tier_totals.get(k, 0)}
            for k, v in sorted(TIER_MAP.items(), key=lambda kv: kv[1].get("order", 99))
        ],
        "leaders": LEADERSHIP_DATA,
        "tree": (build_tree(full_rows, photo_map, PLACEHOLDERS,
                            visible_names={r["name"] for r in rows if r.get("name")})
                 if full_rows else {"nodes": [], "rootNames": [], "children": {}}),
        "pivot": pivot_out,
        "rows": rows,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w") as f:
        json.dump(out, f, indent=2, ensure_ascii=False)

    print(f"wrote {OUT}")
    print(f"  rows: {total}   on RRL: {on_rrl}   eng: {eng} ({out['meta']['totals']['engineeringPct']}%)   product builders: {builders} ({out['meta']['totals']['productBuildersPct']}%)")
    print(f"  photo hits: {photo_hits} / {total}")
    print(f"  facets: {len(out['facets']['project'])} projects, {len(out['facets']['track'])} tracks")
    print(f"  tiers:  " + " · ".join(f"{t['key']}={t['total']}" for t in out["tiers"]))
    leaders = out["leaders"]
    print(f"  leaders: {len(leaders.get('founders', []))} founders, "
          f"{len(leaders.get('cSuite', []))} c-suite, "
          f"{len(leaders.get('projectLeads', {}))} project leads")
    tree = out["tree"]
    print(f"  tree:    {len(tree['nodes'])} nodes (billed-only) · {len(tree['rootNames'])} root(s) {tree['rootNames']}")
    if tree.get("droppedPlaceholder") or tree.get("droppedNonBilled"):
        dp = len(tree.get("droppedPlaceholder", []))
        dn = len(tree.get("droppedNonBilled", []))
        print(f"           dropped: {dp} placeholders + {dn} non-billed (reparented up the chain)")
    enriched = sum(1 for r in rows if r.get("cxo"))
    print(f"  enrich:  {enriched}/{total} rows joined to full roster (cxo + location)")

    # Unmapped report — surfaces xlsx drift / missing mappings.json entries.
    if unmapped_projects or unmapped_tracks or untiered_projects:
        print()
        print("UNMAPPED · add to organisation/mappings.json:")
        for p in sorted(unmapped_projects):
            print(f"  project:  {p!r}")
        for t in sorted(unmapped_tracks):
            print(f"  track:    {t!r}")
        for p in sorted(untiered_projects):
            print(f"  no-tier:  {p!r}  (project mapped but missing 'tier' field)")
    else:
        print("  unmapped: none — every xlsx Project + Track has a mappings.json entry, every project has a tier")
    return 0


if __name__ == "__main__":
    sys.exit(main())
