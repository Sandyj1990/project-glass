"""Read the Fynd Studio Portfolio xlsx and emit a videos.json for the website.

Filters: Reliance umbrella only — explicit excludes are encoded below. Every
remaining row is mapped to a brand via a title-keyword table; rows that don't
match any brand keyword are reported as warnings (and skipped).

Output schema (per video):
    {
      "slug":     "satya-paul-product-video-LzeP1k_qQw0",
      "title":    "Satya Paul Product Video",
      "brand":    "satya-paul",
      "platform": "youtube-long" | "youtube-shorts" | "instagram-reel",
      "videoId":  "LzeP1k_qQw0",      # YouTube id or Instagram shortcode
      "url":      "https://...",
      "embedUrl": "https://www.youtube.com/embed/LzeP1k_qQw0",
      "thumbUrl": "https://i.ytimg.com/vi/LzeP1k_qQw0/maxresdefault.jpg" | null,
      "category": null,                # filled in via overrides later
    }
"""

from __future__ import annotations

import argparse
import json
import logging
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urlparse

import openpyxl

log = logging.getLogger("build_videos_index")


# --------------------------------------------------------------------------- #
# Filtering: titles to exclude entirely. Match is case-insensitive substring.
# --------------------------------------------------------------------------- #
EXCLUDE_TITLE_SUBSTRINGS = [
    # Fynd-self / ambiguous (per Apr 29 review):
    "what makes us humans",  # SOF — explicitly excluded
    "raju video",
    "sof fynd",              # SOF Fynd and Jio Commerce / Create / Forge
    "fynd ai os",            # Fynd's own product
    "ai home tour",
    "ai video magic mirror",
    "hacktimus",             # Internal hackathon / Synapse
    # Non-Reliance brands (excluded per same review):
    "nasher miles",
    "the sleep company",
    "neck pillow by the sleep",
    # Broken source link (Apr 30): IG reel DVp1XXBksXk no longer resolves.
    # Note: "NMAAC Wicked" (YouTube short, different typo) stays in.
    "nmacc wicked video",
]


# --------------------------------------------------------------------------- #
# Brand mapping. Each entry: (regex matched against the title, brand_slug).
# Order matters — first match wins. Specific patterns first.
# --------------------------------------------------------------------------- #
BRAND_RULES: list[tuple[re.Pattern, str]] = [
    (re.compile(r"\breliance\s+digital\b", re.IGNORECASE), "reliance-digital"),
    (re.compile(r"\breliance\s+jewels?\b|reliance\s+rangoli", re.IGNORECASE), "reliance-jewels"),
    (re.compile(r"\bjiomart\b", re.IGNORECASE), "jiomart"),
    # NMACC has at least one typo in the sheet ("NMAAC"), so allow either spelling.
    (re.compile(r"\bnma+c+\b|\bnutcracker\b", re.IGNORECASE), "nmacc"),
    (re.compile(r"\bsatya\s*paul\b", re.IGNORECASE), "satya-paul"),
    (re.compile(r"\bswadesh\b", re.IGNORECASE), "swadesh"),
    (re.compile(r"\bdnmx\b", re.IGNORECASE), "dnmx"),
    (re.compile(r"\bmetro\b", re.IGNORECASE), "metro-wholesale"),
    (re.compile(r"\bsuperdry\b", re.IGNORECASE), "superdry"),
]


# --------------------------------------------------------------------------- #
# URL parsing → (platform, video_id, embed_url, thumb_url)
# --------------------------------------------------------------------------- #
_YT_LONG_HOSTS = {"youtube.com", "www.youtube.com", "m.youtube.com"}
_YT_SHORT_HOSTS = {"youtu.be"}
_IG_HOSTS = {"www.instagram.com", "instagram.com"}


def parse_video_url(url: str) -> dict | None:
    """Return platform + id + embed + thumb for a YouTube/Instagram URL.

    Returns None if the URL isn't recognized.
    """
    try:
        u = urlparse(url.strip())
    except Exception:
        return None
    host = (u.hostname or "").lower()
    path = u.path or ""

    # YouTube Shorts: youtube.com/shorts/<id>
    m = re.match(r"^/shorts/([\w-]{6,})", path)
    if host in _YT_LONG_HOSTS and m:
        vid = m.group(1)
        return {
            "platform": "youtube-shorts",
            "videoId": vid,
            "embedUrl": f"https://www.youtube.com/embed/{vid}",
            "thumbUrl": f"https://i.ytimg.com/vi/{vid}/hqdefault.jpg",
        }

    # YouTube long-form: youtube.com/watch?v=<id>
    if host in _YT_LONG_HOSTS and path == "/watch":
        from urllib.parse import parse_qs
        q = parse_qs(u.query or "")
        if "v" in q and q["v"]:
            vid = q["v"][0]
            return {
                "platform": "youtube-long",
                "videoId": vid,
                "embedUrl": f"https://www.youtube.com/embed/{vid}",
                "thumbUrl": f"https://i.ytimg.com/vi/{vid}/maxresdefault.jpg",
            }

    # YouTube short URL: youtu.be/<id>
    if host in _YT_SHORT_HOSTS:
        m = re.match(r"^/([\w-]{6,})", path)
        if m:
            vid = m.group(1)
            return {
                "platform": "youtube-long",
                "videoId": vid,
                "embedUrl": f"https://www.youtube.com/embed/{vid}",
                "thumbUrl": f"https://i.ytimg.com/vi/{vid}/maxresdefault.jpg",
            }

    # Instagram Reel: instagram.com/reel(s)?/<shortcode>
    m = re.match(r"^/reels?/([\w-]+)", path)
    if host in _IG_HOSTS and m:
        code = m.group(1)
        return {
            "platform": "instagram-reel",
            "videoId": code,
            "embedUrl": f"https://www.instagram.com/reel/{code}/embed/",
            "thumbUrl": None,  # Instagram has no public thumb endpoint
        }

    # Instagram post (non-reel): instagram.com/p/<shortcode>
    m = re.match(r"^/p/([\w-]+)", path)
    if host in _IG_HOSTS and m:
        code = m.group(1)
        return {
            "platform": "instagram-post",
            "videoId": code,
            "embedUrl": f"https://www.instagram.com/p/{code}/embed/",
            "thumbUrl": None,
        }

    return None


def slugify(text: str) -> str:
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-") or "untitled"


def is_excluded(title: str) -> bool:
    t = title.lower()
    return any(s in t for s in EXCLUDE_TITLE_SUBSTRINGS)


def detect_brand(title: str) -> str | None:
    for pattern, slug in BRAND_RULES:
        if pattern.search(title):
            return slug
    return None


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Build data/videos.json from the portfolio xlsx.")
    parser.add_argument(
        "--xlsx",
        default="docs/Fynd Studio - Portfolio.xlsx",
        help="Path to the source Excel file.",
    )
    parser.add_argument(
        "--output",
        default="data/videos.json",
        help="Where to write the generated index.",
    )
    parser.add_argument("--log-level", default="INFO")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=getattr(logging, args.log_level),
        format="%(asctime)s %(levelname)-5s %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        log.error("Source xlsx not found: %s", xlsx_path)
        return 2

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    excluded = []
    unmapped = []
    bad_url = []
    videos: list[dict] = []
    seen_slugs: set[str] = set()

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            continue  # header
        title = (row[0] or "").strip() if len(row) > 0 else ""
        url = (row[1] or "").strip() if len(row) > 1 else ""
        if not title or not url:
            continue

        if is_excluded(title):
            excluded.append((row_idx, title))
            continue

        brand = detect_brand(title)
        if not brand:
            unmapped.append((row_idx, title, url))
            continue

        meta = parse_video_url(url)
        if not meta:
            bad_url.append((row_idx, title, url))
            continue

        # Slug = brand-titlepart-id (id keeps near-duplicates apart, e.g. 4 Satya Paul shorts).
        title_slug = slugify(title)
        # Trim brand prefix from title slug if it's already there.
        if title_slug.startswith(brand + "-"):
            title_slug = title_slug[len(brand) + 1 :] or "video"
        slug = f"{brand}-{title_slug}-{meta['videoId']}"
        # Final fallback if collision.
        n = 1
        base_slug = slug
        while slug in seen_slugs:
            n += 1
            slug = f"{base_slug}-{n}"
        seen_slugs.add(slug)

        videos.append(
            {
                "slug": slug,
                "title": title,
                "brand": brand,
                "platform": meta["platform"],
                "videoId": meta["videoId"],
                "url": url,
                "embedUrl": meta["embedUrl"],
                "thumbUrl": meta["thumbUrl"],
                "category": None,
                "sourceRow": row_idx,
            }
        )

    # Group + counts for the report.
    by_brand: dict[str, list[str]] = {}
    by_platform: dict[str, int] = {}
    for v in videos:
        by_brand.setdefault(v["brand"], []).append(v["slug"])
        by_platform[v["platform"]] = by_platform.get(v["platform"], 0) + 1

    out = {
        "generatedAt": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source": str(xlsx_path),
        "totals": {
            "videos": len(videos),
            "excluded": len(excluded),
            "unmapped": len(unmapped),
            "badUrl": len(bad_url),
            "byPlatform": by_platform,
            "byBrand": {b: len(s) for b, s in sorted(by_brand.items())},
        },
        "videos": sorted(videos, key=lambda v: (v["brand"], v["title"], v["videoId"])),
    }

    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(out, indent=2))

    log.info("Wrote %s", out_path)
    log.info("  videos: %d", len(videos))
    log.info("  excluded: %d", len(excluded))
    log.info("  unmapped (skipped): %d", len(unmapped))
    log.info("  bad url (skipped): %d", len(bad_url))
    log.info("  by brand: %s", out["totals"]["byBrand"])
    log.info("  by platform: %s", out["totals"]["byPlatform"])
    if unmapped:
        log.warning("Unmapped rows (no brand keyword matched):")
        for row_idx, title, url in unmapped:
            log.warning("  row %d: %r → %s", row_idx, title, url)
    if bad_url:
        log.warning("Bad URL rows:")
        for row_idx, title, url in bad_url:
            log.warning("  row %d: %r → %s", row_idx, title, url)
    return 0


if __name__ == "__main__":
    sys.exit(main())
